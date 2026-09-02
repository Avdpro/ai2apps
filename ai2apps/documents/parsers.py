"""Local, dependency-light document parsing adapters."""

from __future__ import annotations

import csv
import io
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True, slots=True)
class ParsedBlock:
    text: str
    kind: str = "text"
    page: int | None = None
    section: str | None = None
    sheet: str | None = None
    slide: int | None = None
    cell_range: str | None = None


class DocumentParserError(RuntimeError):
    pass


class DocumentParser:
    name = "ai2apps-local"
    version = "1"

    def parse(self, path: Path, filename: str, media_type: str) -> list[ParsedBlock]:
        suffix = Path(filename).suffix.lower()
        if suffix in {".html", ".htm"}:
            return self._html(path)
        if suffix in {
            ".py",
            ".js",
            ".ts",
            ".tsx",
            ".jsx",
            ".swift",
            ".rs",
            ".go",
            ".java",
            ".c",
            ".h",
            ".cpp",
            ".hpp",
            ".css",
            ".scss",
            ".sql",
            ".sh",
            ".yaml",
            ".yml",
            ".toml",
        }:
            return self._code(path)
        if suffix in {".csv", ".tsv"} or media_type in {
            "text/csv",
            "text/tab-separated-values",
        }:
            return self._csv(path)
        if suffix in {".txt", ".md", ".json"} or media_type.startswith("text/"):
            return self._text(path)
        if suffix == ".xlsx":
            return self._xlsx(path)
        if suffix == ".pdf":
            blocks = self._pdf(path)
            if blocks:
                return blocks
        if suffix in {".docx", ".pptx", ".pdf"}:
            return self._markitdown(path, suffix)
        raise DocumentParserError(f"Unsupported document type: {suffix or media_type}")

    def _text(self, path: Path) -> list[ParsedBlock]:
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = path.read_text(encoding="utf-8", errors="replace")
        return self._chunk(text)

    def _html(self, path: Path) -> list[ParsedBlock]:
        source = path.read_text(encoding="utf-8", errors="replace")
        try:
            from bs4 import BeautifulSoup

            document = BeautifulSoup(source, "html.parser")
            for node in document(["script", "style", "noscript", "template"]):
                node.decompose()
            text = document.get_text("\n", strip=True)
        except ImportError:
            text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", source)
            text = re.sub(r"(?s)<[^>]+>", "\n", text)
        return self._chunk(text)

    def _code(self, path: Path) -> list[ParsedBlock]:
        text = path.read_text(encoding="utf-8", errors="replace")
        lines = text.splitlines()
        blocks = []
        for start in range(0, len(lines), 160):
            selected = lines[start : start + 160]
            content = "\n".join(selected).strip()
            if not content:
                continue
            section = next(
                (
                    line.strip()[:200]
                    for line in selected
                    if re.match(
                        r"\s*(?:async\s+)?(?:def|class|function|func|fn|interface|struct|enum)\s+",
                        line,
                    )
                ),
                f"lines {start + 1}-{start + len(selected)}",
            )
            blocks.append(ParsedBlock(content, kind="code", section=section))
        return blocks

    def _csv(self, path: Path) -> list[ParsedBlock]:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        return [
            ParsedBlock(
                text=" | ".join(str(value) for value in row),
                kind="table_row",
                cell_range=f"A{index + 1}",
            )
            for index, row in enumerate(rows)
            if any(str(value).strip() for value in row)
        ]

    def _xlsx(self, path: Path) -> list[ParsedBlock]:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise DocumentParserError("XLSX support requires openpyxl") from exc
        workbook = load_workbook(
            io.BytesIO(path.read_bytes()), read_only=True, data_only=True
        )
        blocks: list[ParsedBlock] = []
        try:
            for sheet in workbook.worksheets:
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    values = ["" if value is None else str(value) for value in row]
                    if not any(value.strip() for value in values):
                        continue
                    blocks.append(
                        ParsedBlock(
                            text=" | ".join(values),
                            kind="table_row",
                            sheet=sheet.title,
                            cell_range=f"A{row_index}:{get_column_letter(sheet.max_column)}{row_index}",
                        )
                    )
        finally:
            workbook.close()
        return blocks

    def _pdf(self, path: Path) -> list[ParsedBlock]:
        try:
            import pdfplumber
        except ImportError:
            return []
        blocks: list[ParsedBlock] = []
        with pdfplumber.open(path) as document:
            for page_index, page in enumerate(document.pages, 1):
                text = (page.extract_text() or "").strip()
                for block in self._chunk(text):
                    blocks.append(ParsedBlock(text=block.text, page=page_index))
        return blocks

    def _markitdown(self, path: Path, suffix: str) -> list[ParsedBlock]:
        try:
            from markitdown import MarkItDown
        except ImportError as exc:
            raise DocumentParserError(
                "Office/PDF parsing requires markitdown[pdf,docx,pptx]"
            ) from exc
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix) as temporary:
                with path.open("rb") as source:
                    shutil.copyfileobj(source, temporary)
                temporary.flush()
                result = MarkItDown(enable_plugins=False).convert(temporary.name)
            text = (getattr(result, "text_content", None) or "").strip()
        except Exception as exc:
            raise DocumentParserError(f"Document conversion failed: {exc}") from exc
        if not text:
            raise DocumentParserError(
                "No extractable text found; this document may require OCR"
            )
        chunks = self._chunk(text)
        if suffix == ".pptx":
            return [
                ParsedBlock(text=item.text, slide=index + 1)
                for index, item in enumerate(chunks)
            ]
        return chunks

    @staticmethod
    def _chunk(text: str, limit: int = 4000) -> list[ParsedBlock]:
        paragraphs = [
            item.strip() for item in re.split(r"\n\s*\n", text) if item.strip()
        ]
        blocks: list[ParsedBlock] = []
        for paragraph in paragraphs:
            section = (
                paragraph.splitlines()[0].lstrip("# ")
                if paragraph.startswith("#")
                else None
            )
            for start in range(0, len(paragraph), limit):
                blocks.append(
                    ParsedBlock(paragraph[start : start + limit], section=section)
                )
        return blocks
